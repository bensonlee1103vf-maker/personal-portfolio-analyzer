from datetime import date
from io import BytesIO

import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import streamlit as st
import yfinance as yf

# 設定 matplotlib 字體以支援中文（優先順序：Noto Sans CJK → Microsoft JhengHei → fallback）
# 在 Streamlit Cloud 環境中，透過 packages.txt 安裝 fonts-noto-cjk 以獲得 Noto Sans CJK 字體
plt.rcParams["font.sans-serif"] = [
    "Noto Sans CJK TC",
    "Noto Sans CJK JP",
    "Noto Sans CJK SC",
    "Microsoft JhengHei",
    "Arial Unicode MS",
    "DejaVu Sans",
]
plt.rcParams["axes.unicode_minus"] = False

BACKGROUND_COLOR = "#14232b"
REQUIRED_COLUMNS = ["symbol", "market", "shares", "cost"]
OPTIONAL_COLUMNS = ["price"]
PORTFOLIO_COLUMNS = ["symbol", "market", "category", "shares", "cost"]
COLUMN_ALIASES = {
    "symbol": "symbol",
    "股票代號": "symbol",
    "代號": "symbol",
    "ticker": "symbol",
    "market": "market",
    "市場": "market",
    "shares": "shares",
    "股數": "shares",
    "數量": "shares",
    "持有股數": "shares",
    "cost": "cost",
    "成本": "cost",
    "平均成本": "cost",
    "均價": "cost",
    "買入均價": "cost",
    "category": "category",
    "類別": "category",
    "分類": "category",
    "price": "price",
    "現價": "price",
    "價格": "price",
    "目前價格": "price",
    "市價": "price",
}
CHART_FONT_SCALE = 1.5
CATEGORY_OPTIONS = [
    "未分類",
    "台股ETF",
    "美股ETF",
    "大型科技",
    "AI基礎建設",
    "AI軟體/資料分析",
    "半導體",
    "半導體記憶體",
    "加密貨幣相關",
    "金融科技",
    "醫療健康",
    "電商/消費平台",
    "電動車/未來交通",
    "太空科技",
    "核能/能源科技",
    "無人機/通訊科技",
    "資安",
    "電子零組件",
    "傳產/民生消費",
    "金融股",
    "其他",
]
SAMPLE_CSV = """symbol,market,category,shares,cost
NVDA,US,AI基礎建設,10,150
AAPL,US,大型科技,5,180
QQQ,US,美股ETF,3,500
0050,TW,台股ETF,1000,180
006208,TW,台股ETF,500,110
"""


def create_sample_excel_bytes() -> tuple:
    """建立範例 Excel 的 bytes，供側邊欄下載使用。"""
    df = pd.read_csv(BytesIO(SAMPLE_CSV.encode("utf-8-sig")))
    output = BytesIO()

    # 嘗試多個 engine，避免缺少某個套件時直接 crash
    engines = ["xlsxwriter", "openpyxl", None]
    for engine in engines:
        try:
            if engine:
                with pd.ExcelWriter(output, engine=engine) as writer:
                    df.to_excel(writer, index=False, sheet_name="Sheet1")
            else:
                with pd.ExcelWriter(output) as writer:
                    df.to_excel(writer, index=False, sheet_name="Sheet1")

            return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "sample_portfolio.xlsx"
        except Exception:
            output.seek(0)
            output.truncate(0)
            continue

    # 若所有 engine 都失敗，回傳 CSV bytes 作為 fallback，並提醒使用者安裝套件
    try:
        st.warning(
            "無法建立範例 Excel，已改以 CSV 格式提供下載。若要產生 Excel，請安裝 xlsxwriter 或 openpyxl 套件。"
        )
    except Exception:
        pass

    return SAMPLE_CSV.encode("utf-8-sig"), "text/csv", "sample_portfolio.csv"

MARKET_NAME_MAP = {"US": "美股", "TW": "台股"}
MARKET_CURRENCY_MAP = {"US": "USD", "TW": "TWD"}
DISPLAY_COLUMNS = [
    "symbol",
    "market",
    "category",
    "shares",
    "currency",
    "cost",
    "price",
    "market_value_twd",
    "profit",
    "profit_twd",
    "return_rate",
    "weight",
    "price_status",
]
DISPLAY_COLUMN_NAMES = {
    "symbol": "代號",
    "market": "市場",
    "category": "類別",
    "shares": "股數",
    "currency": "原始幣別",
    "cost": "原始均價",
    "price": "原始現價",
    "market_value_twd": "台幣現值",
    "profit": "原始損益",
    "profit_twd": "台幣損益",
    "return_rate": "報酬率",
    "weight": "持倉比例",
    "price_status": "價格狀態",
}
EXCEL_HOLDINGS_COLUMNS = [
    "symbol",
    "market",
    "category",
    "shares",
    "currency",
    "cost",
    "price",
    "profit",
    "cost_twd",
    "price_twd",
    "cost_value_twd",
    "market_value_twd",
    "profit_twd",
    "return_rate",
    "weight",
    "price_status",
]
EXCEL_HOLDINGS_COLUMN_NAMES = {
    "symbol": "代號",
    "market": "市場",
    "category": "類別",
    "shares": "股數",
    "currency": "原始幣別",
    "cost": "原始均價",
    "price": "原始現價",
    "profit": "原始損益",
    "cost_twd": "台幣均價",
    "price_twd": "台幣現價",
    "cost_value_twd": "台幣成本",
    "market_value_twd": "台幣市值",
    "profit_twd": "台幣損益",
    "return_rate": "報酬率",
    "weight": "持倉比例",
    "price_status": "價格狀態",
}
EXCEL_SHEET_NAMES = {
    "summary": "Summary",
    "holdings": "Holdings",
    "market": "Market Allocation",
    "category": "Category Allocation",
    "risk": "Risk Flags",
}


def load_portfolio(uploaded_file, debug_mode=False):
    """相容用的讀取函式，維持原有介面（內部已改為支援 Excel）。"""
    return load_uploaded_file(uploaded_file, debug_mode=debug_mode)


def is_instruction_sheet(sheet_name: str) -> bool:
    """判斷工作表名稱是否為說明／README 類型的頁籤。"""
    if not isinstance(sheet_name, str):
        return False

    normalized = sheet_name.strip().lower()
    instruction_keywords = [
        "說明",
        "readme",
        "instruction",
        "instructions",
        "help",
        "note",
    ]
    return any(keyword in normalized for keyword in instruction_keywords)


def load_uploaded_file(uploaded_file, debug_mode=False):
    """讀取使用者上傳的 Excel（.xlsx/.xls）。

    - 先讀取整張工作表為 header=None 的 raw_df。
    - 自動偵測最像欄位名稱的列，並從該列下一列開始讀資料。
    - 讀取後呼叫 `clean_uploaded_portfolio` 做進一步檢查與轉型。
    """
    # 檔名副檔名檢查
    filename = getattr(uploaded_file, "name", "")
    ext = filename.split(".")[-1].lower() if filename and "." in filename else ""
    if ext not in {"xlsx", "xls"}:
        st.error("只支援 Excel 檔 (.xlsx, .xls)。")
        st.stop()

    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception:
        st.error("無法讀取上傳的 Excel 檔案。請確認檔案不是損毀或受保護的格式。")
        st.stop()

    sheets = xls.sheet_names
    candidate_sheets = [sheet for sheet in sheets if not is_instruction_sheet(sheet)]
    sheet = candidate_sheets[0] if candidate_sheets else sheets[0]

    try:
        raw_df = pd.read_excel(
            xls,
            sheet_name=sheet,
            header=None,
            dtype=str,
            keep_default_na=False,
        )
    except Exception:
        st.error("讀取工作表失敗，請確認 Excel 內容為有效表格。")
        st.stop()

    header_index = detect_header_row(raw_df)
    if header_index is None:
        st.error(
            "找不到欄位名稱列。請確認 Excel 中有股票代號、市場、股數、平均成本等欄位。"
        )
        st.stop()

    raw_headers = raw_df.iloc[header_index].astype(str).str.strip()
    df = raw_df.iloc[header_index + 1 :].copy()
    df.columns = raw_headers
    df = df.reset_index(drop=True)

    df.columns = df.columns.astype(str).str.strip()
    df = normalize_uploaded_columns(df)
    if debug_mode:
        show_header_debug_info(header_index, raw_headers.tolist(), df.columns.tolist())

    inspection = inspect_portfolio_data(df)

    if inspection["missing_columns"]:
        st.error(
            "必要欄位缺少：" + ", ".join(inspection["missing_columns"]) + "。"
        )
        st.stop()

    cleaned = clean_uploaded_portfolio(df)
    return cleaned, inspection


def load_sample_portfolio():
    """載入內建範例資料，供使用者試玩用。"""
    df = pd.read_csv(BytesIO(SAMPLE_CSV.encode("utf-8-sig")), dtype=str)
    df.columns = df.columns.str.strip()
    df = normalize_uploaded_columns(df)
    inspection = inspect_portfolio_data(df)
    cleaned = clean_uploaded_portfolio(df)
    return cleaned, inspection


def normalize_uploaded_columns(df: pd.DataFrame) -> pd.DataFrame:
    """將中文欄位名對應成內部使用的英文欄位名。"""
    df = df.copy()
    normalized_columns = []
    for col in df.columns:
        key = str(col).strip().lower()
        normalized_columns.append(COLUMN_ALIASES.get(key, str(col).strip()))
    df.columns = normalized_columns

    if df.columns.duplicated().any():
        duplicated = df.columns[df.columns.duplicated()].unique().tolist()
        st.warning(
            "上傳的欄位名稱中有重複對應，系統將使用第一個出現的欄位，其他重複欄位會被忽略。"
        )
        for label in duplicated:
            cols = [col for col in df.columns if col == label]
            df[label] = df.loc[:, cols].iloc[:, 0]
        df = df.loc[:, ~df.columns.duplicated(keep="first")]

    return df


def detect_header_row(raw_df: pd.DataFrame):
    """自動偵測 Excel 中最像欄位名稱的列。"""
    if raw_df.empty:
        return None

    required_fields = set(REQUIRED_COLUMNS)
    normalized_aliases = {key.lower(): value for key, value in COLUMN_ALIASES.items()}

    for row_index, row in raw_df.iterrows():
        matched_fields = set()
        for cell in row:
            cell_value = str(cell).strip().lower()
            if not cell_value or cell_value == "nan":
                continue

            canonical_name = normalized_aliases.get(cell_value)
            if canonical_name in required_fields:
                matched_fields.add(canonical_name)

        if len(matched_fields) >= 3:
            return row_index

    return None


def show_header_debug_info(header_index: int, raw_headers, normalized_headers):
    """顯示 header row 偵測的除錯資訊。"""
    st.info(
        f"偵測到的欄位名稱列：第 {header_index + 1} 列"
    )
    st.write("原始欄位名稱：", list(raw_headers))
    st.write("normalize 後的欄位名稱：", list(normalized_headers))


def inspect_portfolio_data(df: pd.DataFrame) -> dict:
    """檢查資料欄位與基本格式，回傳檢查結果供畫面顯示。"""
    checks = {
        "missing_columns": [],
        "invalid_market": [],
        "shares_errors": None,
        "cost_errors": None,
        "price_errors": None,
        "has_unclassified": False,
        "price_present": "price" in df.columns,
    }

    # 檢查是否存在必要欄位
    checks["missing_columns"] = [col for col in REQUIRED_COLUMNS if col not in df.columns]

    # 檢查 market
    if "market" in df.columns:
        markets = df["market"].astype(str).str.strip().str.upper()
        checks["invalid_market"] = sorted(set(markets.unique()) - {"US", "TW"})

    def numeric_check(series: pd.Series) -> dict:
        original = series.astype(str).fillna("").str.strip()
        cleaned = original.str.replace(r"[^0-9.\-]", "", regex=True)
        numeric = pd.to_numeric(cleaned.replace("", pd.NA), errors="coerce")
        bad = numeric.isna() & ~original.isin(["", "nan", "None"])
        empty = original.isin(["", "nan", "None"])
        return {
            "total": len(series),
            "valid": int((~bad & ~empty).sum()),
            "empty": int(empty.sum()),
            "invalid": int(bad.sum()),
            "samples": original[bad].unique().tolist()[:5],
        }

    if "shares" in df.columns:
        checks["shares_errors"] = numeric_check(df["shares"])
    if "cost" in df.columns:
        checks["cost_errors"] = numeric_check(df["cost"])
    if checks["price_present"]:
        checks["price_errors"] = numeric_check(df["price"])

    if "category" not in df.columns or clean_category_series(df["category"]).eq("未分類").any():
        checks["has_unclassified"] = True

    return checks


def show_data_check_results(inspection: dict) -> bool:
    """顯示資料檢查結果，並回傳是否通過基本檢查。"""
    has_blocking_errors = any(
        [
            inspection["missing_columns"],
            inspection["invalid_market"],
            inspection["shares_errors"]
            and (inspection["shares_errors"]["empty"] > 0 or inspection["shares_errors"]["invalid"] > 0),
            inspection["cost_errors"]
            and (inspection["cost_errors"]["empty"] > 0 or inspection["cost_errors"]["invalid"] > 0),
        ]
    )
    has_price_warning = (
        inspection["price_present"]
        and inspection["price_errors"]
        and inspection["price_errors"]["invalid"] > 0
    )

    # 資料完全正常時不顯示檢查區塊，讓主畫面更乾淨。
    if not has_blocking_errors and not has_price_warning:
        return True

    st.subheader("資料檢查結果")

    if inspection["missing_columns"]:
        st.error(
            "必要欄位缺少：" + ", ".join(inspection["missing_columns"]) + "。"
        )

    if inspection["invalid_market"]:
        st.error(
            "market 欄位只能是 US 或 TW，目前偵測到："
            + ", ".join(inspection["invalid_market"]) + "。"
        )

    def render_numeric_check(name: str, result: dict, required: bool = True):
        if result is None:
            return
        if result["empty"] > 0 or result["invalid"] > 0:
            if required:
                st.error(
                    f"{name} 欄位: {result['empty']} 筆空白，{result['invalid']} 筆格式不正確。"
                )
            else:
                st.warning(
                    f"{name} 欄位：{result['invalid']} 筆格式不正確，會視為無效價格。"
                )

    render_numeric_check("shares", inspection["shares_errors"])
    render_numeric_check("cost", inspection["cost_errors"])
    if inspection["price_present"]:
        render_numeric_check("price", inspection["price_errors"], required=False)
        if inspection["price_errors"] and inspection["price_errors"]["invalid"] > 0:
            st.info("price 欄位僅在勾選使用 Excel price 時有效，格式錯誤的資料將被忽略。")

    return not has_blocking_errors


def render_portfolio_health(report):
    """顯示投資組合健康檢查提醒。"""
    st.subheader("投資組合健康檢查")
    st.write("以下提醒為中性資訊，僅供參考，不構成投資建議。")

    total_value = report["market_value_twd"].sum()
    if total_value == 0:
        st.info("尚未有可計算的持倉市值。")
        return

    messages = []
    largest_weight = report["weight"].max()
    if largest_weight > 0.3:
        symbol = report.loc[report["weight"].idxmax(), "symbol"]
        messages.append(
            f"單一持股 {symbol} 佔比 {format_percent(largest_weight)}，高於 30%。"
        )
    else:
        messages.append(f"單一持股最大佔比為 {format_percent(largest_weight)}。")

    category_weights = report.groupby("category")["market_value_twd"].sum() / total_value
    top_category = category_weights.idxmax()
    top_category_weight = category_weights.max()
    if top_category_weight > 0.5:
        messages.append(
            f"分類 {top_category} 佔比 {format_percent(top_category_weight)}，高於 50%。"
        )
    else:
        messages.append(
            f"最大的分類 {top_category} 佔比為 {format_percent(top_category_weight)}。"
        )

    unclassified_ratio = category_weights.get("未分類", 0)
    if unclassified_ratio > 0.1:
        messages.append(
            f"未分類持股佔比 {format_percent(unclassified_ratio)}，建議補上分類以提升分析準確度。"
        )
    else:
        messages.append(
            f"未分類持股佔比為 {format_percent(unclassified_ratio)}。"
        )

    market_weights = report.groupby("market")["market_value_twd"].sum() / total_value
    us_ratio = market_weights.get("US", 0)
    tw_ratio = market_weights.get("TW", 0)
    if us_ratio > 0.8:
        messages.append(f"美股佔比 {format_percent(us_ratio)}，超過 80%。")
    elif tw_ratio > 0.8:
        messages.append(f"台股佔比 {format_percent(tw_ratio)}，超過 80%。")
    else:
        messages.append(
            f"美股佔比 {format_percent(us_ratio)}，台股佔比 {format_percent(tw_ratio)}。"
        )

    etf_weight = report[report["category"].astype(str).str.contains("ETF")]["market_value_twd"].sum() / total_value
    messages.append(f"ETF 類別佔比為 {format_percent(etf_weight)}。")

    for message in messages:
        st.write(f"- {message}")


def clean_uploaded_portfolio(df: pd.DataFrame) -> pd.DataFrame:
    """在讀取 Excel 後呼叫，統一清理欄位名稱與內容，並做欄位型別驗證。"""

    # 檢查必要欄位名稱是否存在（先不用轉型）
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        # 顯示通用的使用者友善訊息
        msg = (
            "Excel 缺少必要欄位。必要欄位為 symbol, market, shares, cost。category 可選填。"
        )
        st.error(msg)
        raise ValueError(msg)

    result = df.copy()

    # 文字欄位先做 strip
    for col in result.columns:
        if result[col].dtype == object or pd.api.types.is_string_dtype(result[col]):
            result[col] = result[col].astype(str).str.strip()

    # symbol 強制字串，保留前置 0
    result["symbol"] = result["symbol"].astype(str).str.strip()

    # category 若不存在或為空就補成 未分類
    if "category" not in result.columns:
        result["category"] = "未分類"
    result["category"] = clean_category_series(result["category"])

    # market 標準化成大寫
    result["market"] = result["market"].astype(str).str.strip().str.upper()

    # 檢查 market 是否只有 US / TW
    invalid_market = sorted(set(result["market"].unique()) - {"US", "TW"})
    if invalid_market:
        msg = "market 只能填 US 或 TW。"
        st.error(msg)
        raise ValueError(msg)

    # 解析 numeric 欄位的 helper
    def parse_numeric_series(series: pd.Series, col_name: str, required: bool = True) -> pd.Series:
        original = series.astype(str).str.strip()

        # 移除千分位逗號與貨幣符號，保留負號與小數點
        cleaned = original.str.replace(r"[^0-9.\-]", "", regex=True)

        # 轉成數字，失敗的會成為 NaN
        numeric = pd.to_numeric(cleaned.replace("", pd.NA), errors="coerce")

        # 找出空白(使用者未填)與轉換失敗的列，供友善訊息顯示
        empty_rows = original[original.isin(["", "nan", "None"])].index.tolist()
        failed_rows = numeric[numeric.isna() & ~original.isin(["", "nan", "None"])].index.tolist()

        messages = []
        if empty_rows and required:
            messages.append(f"欄位 {col_name} 有 {len(empty_rows)} 筆空白值，請補上後再上傳。")
        if failed_rows:
            sample_vals = original.loc[failed_rows].unique()[:5]
            if required:
                messages.append(
                    f"欄位 {col_name} 有 {len(failed_rows)} 筆格式不正確（範例：{', '.join(map(str, sample_vals))}），請使用數字或像 1,000 / NT$1,000 / $100 的格式。"
                )
            else:
                st.warning(
                    f"欄位 {col_name} 有 {len(failed_rows)} 筆格式不正確（範例：{', '.join(map(str, sample_vals))}），這些值將視為無效。"
                )

        if messages:
            # 先在介面上顯示具體錯誤，再以通用錯誤停止流程
            for m in messages:
                st.error(m)
            if required:
                raise ValueError(
                    "Excel 格式有問題，請檢查欄位與數字格式，或下載範例 Excel 重新填寫。"
                )

        return numeric.astype(float)

    # 轉換 shares 與 cost
    result["shares"] = parse_numeric_series(result["shares"], "shares")
    result["cost"] = parse_numeric_series(result["cost"], "cost")

    if "price" in result.columns:
        # 可接受 Excel 中的 price 欄位為手動現價或備用價格。
        result["price"] = parse_numeric_series(result["price"], "price", required=False)

    # 只回傳必要欄位與可選價格欄位，方便後續選用 Excel 價格。
    optional = [col for col in OPTIONAL_COLUMNS if col in result.columns]
    return result[PORTFOLIO_COLUMNS + optional].copy()


def ensure_category_column(df):
    """確保資料一定有 category 欄位，空白分類會補成「未分類」。"""
    result = df.copy()

    # 如果使用者的上傳檔案沒有 category 欄位，就直接新增一欄。
    if "category" not in result.columns:
        result["category"] = "未分類"
        return result

    result["category"] = clean_category_series(result["category"])

    return result


def clean_category_series(series):
    """把 category 欄位中的空值、None、空字串統一整理成「未分類」。"""
    cleaned = series.fillna("").astype(str).str.strip()
    return cleaned.mask(cleaned == "", "未分類")


def get_category_options(df):
    """建立下拉選單選項，保留上傳檔案原本自訂的分類名稱。"""
    options = CATEGORY_OPTIONS.copy()

    for category in df["category"].dropna().astype(str).str.strip().unique():
        if category and category not in options:
            options.append(category)

    return options


def update_category_options(df, custom_category):
    """合併固定分類、上傳檔案自帶分類與使用者新增的自訂分類。"""
    options = get_category_options(df)
    custom_category = custom_category.strip() if custom_category else ""

    # 使用者新增的分類只加入選單，不會覆蓋原本資料。
    if custom_category and custom_category not in options:
        options.append(custom_category)

    return options


def to_yahoo_symbol(row):
    """依 market 轉成 Yahoo Finance 使用的代號。"""
    if row["market"] == "TW":
        return f"{row['symbol'].zfill(4)}.TW"
    return row["symbol"]


@st.cache_data(ttl=900, show_spinner=False)
def get_latest_price(yahoo_symbol):
    """使用 yfinance 抓最新收盤價。"""
    try:
        stock = yf.Ticker(yahoo_symbol)
        data = stock.history(period="5d")
        close_prices = data["Close"].dropna() if not data.empty else pd.Series(dtype=float)
    except Exception:
        return None

    if close_prices.empty:
        return None

    return float(close_prices.iloc[-1])


@st.cache_data(ttl=900, show_spinner=False)
def get_usd_twd_rate():
    """抓 USD/TWD 匯率，作為美股換算台幣的依據。"""
    fx_data = yf.Ticker("USDTWD=X").history(period="5d")
    close_prices = fx_data["Close"].dropna() if not fx_data.empty else pd.Series(dtype=float)

    if close_prices.empty:
        raise ValueError("無法取得 USD/TWD 匯率。")

    return float(close_prices.iloc[-1])


def calculate_portfolio(df, usd_twd, use_excel_price=False):
    """計算價格、台幣市值、損益、報酬率與權重。"""
    result = df.copy()

    result["currency"] = result["market"].map(MARKET_CURRENCY_MAP)
    result["yahoo_symbol"] = result.apply(to_yahoo_symbol, axis=1)
    result["yahoo_price"] = result["yahoo_symbol"].apply(get_latest_price)

    if "price" in result.columns:
        result["price"] = pd.to_numeric(result["price"], errors="coerce")
    else:
        result["price"] = pd.Series([pd.NA] * len(result), index=result.index)

    if use_excel_price:
        original_price = result["price"].copy()
        result["price"] = result["price"].fillna(result["yahoo_price"])
        result["price_status"] = original_price.apply(
            lambda value: "使用 Excel price" if pd.notna(value) else "價格正常"
        )
        result.loc[result["price"].isna(), "price_status"] = "價格抓取失敗，暫用成本估算"
    else:
        result["price"] = result["yahoo_price"]
        result["price_status"] = result["price"].apply(
            lambda price: "價格正常" if pd.notna(price) else "價格抓取失敗，暫用成本估算"
        )

    result["price_for_calc"] = result["price"].fillna(result["cost"])

    # 美股價格與成本用 USD/TWD 換成台幣，台股本來就是台幣。
    is_us = result["market"] == "US"
    result["price_twd"] = result["price_for_calc"]
    result.loc[is_us, "price_twd"] = result.loc[is_us, "price_for_calc"] * usd_twd

    result["cost_twd"] = result["cost"]
    result.loc[is_us, "cost_twd"] = result.loc[is_us, "cost"] * usd_twd

    # 原始損益保留各市場原本的幣別：美股是美元，台股是台幣。
    result["profit"] = result["shares"] * (result["price_for_calc"] - result["cost"])
    result["market_value_twd"] = result["shares"] * result["price_twd"]
    result["cost_value_twd"] = result["shares"] * result["cost_twd"]
    result["profit_twd"] = result["market_value_twd"] - result["cost_value_twd"]
    result["return_rate"] = result["profit_twd"] / result["cost_value_twd"]
    result["weight"] = result["market_value_twd"] / result["market_value_twd"].sum()

    return result.sort_values("market_value_twd", ascending=False).reset_index(drop=True)


def make_top_n_summary(data, name_col, value_col, top_n):
    """取前 top_n 大項目，其餘合併成「其他」。
    滑桿選幾大，就真的顯示幾大；剩下的合併成「其他」。
    """
    if data.empty:
        return pd.DataFrame(columns=[name_col, value_col, "weight"])

    sorted_data = data.sort_values(value_col, ascending=False).copy()
    top_items = sorted_data.head(top_n).copy()
    other_items = sorted_data.iloc[top_n:].copy()

    summary = top_items[[name_col, value_col]].copy()

    if not other_items.empty:
        other_row = pd.DataFrame(
            [{name_col: "其他", value_col: other_items[value_col].sum()}]
        )
        summary = pd.concat([summary, other_row], ignore_index=True)

    summary["weight"] = summary[value_col] / summary[value_col].sum()
    return summary


def create_symbol_color_map(df):
    """建立股票代號與顏色的對照表，讓不同圖中的同一支股票顏色一致。"""
    sorted_symbols = (
        df.sort_values("market_value_twd", ascending=False)["symbol"]
        .drop_duplicates()
        .tolist()
    )

    # tab20 顏色清楚、數量也夠多；超過 20 支時會循環使用。
    palette = list(plt.cm.tab20.colors)
    color_map = {
        symbol: palette[index % len(palette)]
        for index, symbol in enumerate(sorted_symbols)
    }
    color_map["其他"] = "#8a8a8a"

    return color_map


def plot_donut_on_ax(
    ax,
    summary,
    name_col,
    value_col,
    title,
    center_text,
    label_size=10,
    pct_size=9,
    center_size=16,
    title_size=14,
    title_y=1.04,
    radius=0.88,
    color_map=None,
):
    """在指定的 matplotlib ax 上畫一張 donut chart。"""
    ax.set_facecolor(BACKGROUND_COLOR)

    label_size *= CHART_FONT_SCALE
    pct_size *= 1.1
    center_size *= CHART_FONT_SCALE
    title_size *= CHART_FONT_SCALE

    if summary.empty or summary[value_col].sum() == 0:
        ax.text(
            0,
            0,
            "無資料",
            ha="center",
            va="center",
            color="white",
            fontsize=13 * CHART_FONT_SCALE,
        )
        ax.set_title(title, fontsize=title_size, color="white", fontweight="bold", y=title_y)
        ax.axis("off")
        return

    colors = None
    if color_map is not None:
        colors = [color_map.get(name, "#9aa4ad") for name in summary[name_col]]

    wedges, texts, autotexts = ax.pie(
        summary[value_col],
        labels=None,
        colors=colors,
        autopct="%1.1f%%",
        startangle=90,
        pctdistance=0.78,
        labeldistance=1.05,
        radius=radius,
        wedgeprops={
            "width": 0.38,
            "edgecolor": BACKGROUND_COLOR,
            "linewidth": 1,
        },
        textprops={
            "color": "white",
            "fontsize": label_size,
        },
    )

    for autotext in autotexts:
        autotext.set_color("white")
        autotext.set_fontsize(pct_size)

    legend_title = "代號" if name_col == "symbol" else "類別"
    legend = ax.legend(
        wedges,
        summary[name_col],
        title=legend_title,
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        fontsize=9,
        frameon=False,
    )
    if legend:
        legend.get_title().set_color("white")
        for text in legend.get_texts():
            text.set_color("white")

    ax.text(
        0,
        0,
        center_text,
        ha="center",
        va="center",
        fontsize=center_size,
        color="white",
        fontweight="bold",
    )

    ax.set_title(
        title,
        fontsize=title_size,
        color="white",
        fontweight="bold",
        y=title_y,
    )
    ax.axis("equal")


def plot_top_holdings_bar(df, top_n, value_col="weight"):
    """畫出前 top_n 大持股的橫向長條圖。"""
    fig, ax = plt.subplots(figsize=(10, max(5, top_n * 0.28)))
    fig.patch.set_facecolor(BACKGROUND_COLOR)
    ax.set_facecolor(BACKGROUND_COLOR)

    top_df = df.sort_values(value_col, ascending=False).head(top_n).copy()
    if top_df.empty:
        ax.text(
            0.5,
            0.5,
            "無持股資料",
            color="white",
            ha="center",
            va="center",
            fontsize=13 * CHART_FONT_SCALE,
            transform=ax.transAxes,
        )
        ax.axis("off")
        return fig

    is_profit_chart = value_col in {"profit", "profit_twd"}
    if is_profit_chart:
        bar_colors = [
            "#00B050" if value > 0 else "#FF0000" if value < 0 else "#9aa4ad"
            for value in top_df[value_col]
        ]
    else:
        bar_colors = ["#4cc9f0" if m == "TW" else "#ff595e" for m in top_df["market"]]

    ax.barh(
        y=top_df["symbol"],
        width=top_df[value_col],
        color=bar_colors,
        edgecolor="white",
        height=0.6,
    )
    ax.invert_yaxis()

    max_abs_value = top_df[value_col].abs().max()
    label_offset = max_abs_value * 0.02 if is_profit_chart and max_abs_value > 0 else 0.005

    for i, (_, row) in enumerate(top_df.iterrows()):
        value = row[value_col]
        text_x = value + label_offset if not is_profit_chart or value >= 0 else value - label_offset
        text_align = "left" if not is_profit_chart or value >= 0 else "right"
        ax.text(
            text_x,
            i,
            format_percent(value) if value_col == "weight" else format_twd(value),
            va="center",
            ha=text_align,
            color="white",
            fontsize=10,
        )

    ax.set_title(
        f"Top {len(top_df)} 持股比例",
        fontsize=16 * CHART_FONT_SCALE,
        color="white",
        fontweight="bold",
    )
    ax.set_xlabel("持倉比例" if value_col == "weight" else "台幣市值", color="white")
    ax.set_ylabel("股票代號", color="white")
    ax.tick_params(axis="x", colors="white")
    ax.tick_params(axis="y", colors="white")

    if is_profit_chart:
        ax.axvline(0, color="white", linewidth=1, alpha=0.6)
        min_value = min(0, top_df[value_col].min() * 1.15)
        max_value = max(0, top_df[value_col].max() * 1.15)
        if min_value == max_value:
            min_value, max_value = -1, 1
        ax.set_xlim(min_value, max_value)
    else:
        ax.set_xlim(0, min(1.0, top_df[value_col].max() * 1.15))
    for spine in ax.spines.values():
        spine.set_color("#555555")

    return fig


def plot_market_bar_on_ax(ax, market_summary):
    """在 dashboard 底部畫出台股與美股的市場比例長條圖。"""
    ax.set_facecolor(BACKGROUND_COLOR)

    if market_summary.empty or market_summary["market_value_twd"].sum() == 0:
        ax.text(
            0.5,
            0.5,
            "無市場資料",
            ha="center",
            va="center",
            color="white",
            fontsize=13 * CHART_FONT_SCALE,
            transform=ax.transAxes,
        )
        ax.axis("off")
        return

    summary = market_summary.copy()
    summary["weight"] = summary["market_value_twd"] / summary["market_value_twd"].sum()

    # 固定市場顯示順序，讓長條圖每次都容易比較。
    market_order = ["TW", "US"]
    summary["market"] = pd.Categorical(summary["market"], categories=market_order, ordered=True)
    summary = summary.sort_values("market")

    colors = {"TW": "#4cc9f0", "US": "#ff595e"}
    left = 0

    for _, row in summary.iterrows():
        width = row["weight"]
        market = row["market"]
        market_name = row["market_name"]

        ax.barh(
            y=0,
            width=width,
            left=left,
            height=0.42,
            color=colors.get(market, "#9aa4ad"),
            edgecolor=BACKGROUND_COLOR,
            linewidth=2,
        )

        if width >= 0.06:
            ax.text(
                left + width / 2,
                0,
                f"{market_name} {format_percent(width)}",
                ha="center",
                va="center",
                color="white",
                fontsize=11 * CHART_FONT_SCALE,
                fontweight="bold",
            )

        left += width

    ax.set_xlim(0, 1)
    ax.set_ylim(-0.5, 0.5)
    ax.set_title(
        "台股 vs 美股比例",
        fontsize=13 * CHART_FONT_SCALE,
        color="white",
        fontweight="bold",
        pad=24,
    )
    ax.set_xticks([])
    ax.set_yticks([])

    for spine in ax.spines.values():
        spine.set_visible(False)


def plot_dashboard(df, top_n):
    """建立左邊總覽、右邊台股/美股、底部市場比例的 dashboard。"""
    # matplotlib 字體設定已在模組初始化時設定（全域配置）
    symbol_color_map = create_symbol_color_map(df)

    overview_summary = make_top_n_summary(
        data=df,
        name_col="symbol",
        value_col="market_value_twd",
        top_n=top_n,
    )

    market_summary = (
        df.groupby("market", as_index=False)["market_value_twd"]
        .sum()
        .sort_values("market_value_twd", ascending=False)
    )
    market_summary["market_name"] = market_summary["market"].map(MARKET_NAME_MAP)

    tw_summary = make_top_n_summary(
        data=df[df["market"] == "TW"],
        name_col="symbol",
        value_col="market_value_twd",
        top_n=top_n,
    )

    us_summary = make_top_n_summary(
        data=df[df["market"] == "US"],
        name_col="symbol",
        value_col="market_value_twd",
        top_n=top_n,
    )

    fig = plt.figure(figsize=(15, 14))
    fig.patch.set_facecolor(BACKGROUND_COLOR)

    # 左邊放整體總覽，右邊台股與美股平分空間，底部放市場比例長條圖。
    gs = fig.add_gridspec(
        3,
        2,
        width_ratios=[1.15, 1],
        height_ratios=[1, 1, 0.46],
        left=0.06,
        right=0.96,
        top=0.82,
        bottom=0.10,
        wspace=0.22,
        hspace=0.72,
    )

    ax_overview = fig.add_subplot(gs[0:2, 0])
    ax_tw = fig.add_subplot(gs[0, 1])
    ax_us = fig.add_subplot(gs[1, 1])
    ax_market = fig.add_subplot(gs[2, :])

    plot_donut_on_ax(
        ax=ax_overview,
        summary=overview_summary,
        name_col="symbol",
        value_col="market_value_twd",
        title=f"投資組合前 {top_n} 大 + 其他",
        center_text="總覽",
        label_size=11,
        pct_size=10,
        center_size=18,
        title_size=15,
        title_y=1.03,
        color_map=symbol_color_map,
    )

    plot_donut_on_ax(
        ax=ax_tw,
        summary=tw_summary,
        name_col="symbol",
        value_col="market_value_twd",
        title=f"台股前 {top_n} 大 + 其他",
        center_text="台股",
        label_size=10,
        pct_size=9,
        center_size=14,
        title_size=13,
        title_y=1.18,
        radius=0.78,
        color_map=symbol_color_map,
    )

    plot_donut_on_ax(
        ax=ax_us,
        summary=us_summary,
        name_col="symbol",
        value_col="market_value_twd",
        title=f"美股前 {top_n} 大 + 其他",
        center_text="美股",
        label_size=10,
        pct_size=9,
        center_size=14,
        title_size=13,
        title_y=1.18,
        radius=0.78,
        color_map=symbol_color_map,
    )

    plot_market_bar_on_ax(ax_market, market_summary)

    fig.suptitle(
        "投資組合總覽",
        fontsize=24 * CHART_FONT_SCALE,
        color="white",
        fontweight="bold",
        y=0.94,
    )

    return fig


def create_summary_sheet(summary):
    """把摘要卡片資料整理成 Excel 的 Summary sheet。"""
    return pd.DataFrame(
        [
            {"metric": "總市值", "value": summary["total_value"]},
            {"metric": "總成本", "value": summary["total_cost"]},
            {"metric": "總損益", "value": summary["total_profit"]},
            {"metric": "總報酬率", "value": summary["total_return"]},
            {"metric": "最大持股", "value": summary["largest_holding"]},
            {"metric": "美股佔比", "value": summary["us_weight"]},
            {"metric": "台股佔比", "value": summary["tw_weight"]},
        ]
    )


def create_excel_holdings_sheet(holdings_df):
    """建立 Excel 用的完整中文持倉明細。"""
    export_df = holdings_df.sort_values("weight", ascending=False).reset_index(drop=True).copy()
    export_df["market"] = export_df["market"].map(MARKET_NAME_MAP)
    export_df = export_df.drop(columns=["index", "level_0"], errors="ignore")
    export_df = export_df[EXCEL_HOLDINGS_COLUMNS].copy()
    return export_df.rename(columns=EXCEL_HOLDINGS_COLUMN_NAMES)


def excel_column_letter(col_idx):
    """把 0-based 欄位位置轉成 Excel 欄名，例如 0 -> A。"""
    letters = ""
    col_num = col_idx + 1

    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        letters = chr(65 + remainder) + letters

    return letters


def create_multi_sheet_excel(
    holdings_df,
    summary,
    market_allocation,
    category_allocation,
    risk_flags,
):
    """建立包含多個工作表的 Excel 報表。"""
    output = BytesIO()
    sheets = {
        EXCEL_SHEET_NAMES["holdings"]: create_excel_holdings_sheet(holdings_df),
        EXCEL_SHEET_NAMES["summary"]: create_summary_sheet(summary),
        EXCEL_SHEET_NAMES["market"]: market_allocation,
        EXCEL_SHEET_NAMES["category"]: category_allocation,
        EXCEL_SHEET_NAMES["risk"]: risk_flags,
    }

    writer = pd.ExcelWriter(output, engine="openpyxl")

    with writer:
        write_excel_sheets(writer, sheets)
        apply_excel_formatting(writer, sheets)

    return output.getvalue()


def write_excel_sheets(writer, sheets):
    """把每個 DataFrame 寫入指定的 Excel sheet。"""
    for sheet_name, sheet_df in sheets.items():
        sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)


def apply_excel_formatting(writer, sheets):
    """在所有 sheets 寫入完成後，統一套用 Excel 格式。"""
    workbook = writer.book

    for sheet_name, sheet_df in sheets.items():
        ws = workbook[sheet_name]
        add_excel_table_and_filter(ws, sheet_name)
        format_worksheet_header(ws)
        apply_number_formats(ws)
        apply_profit_loss_colors(ws)
        highlight_price_status_rows(ws)
        auto_adjust_column_widths(ws)
        ws.freeze_panes = "A2"


def get_header_map(ws):
    """回傳 {欄位名稱: 欄號}，方便依欄位名稱套格式。"""
    return {
        str(cell.value): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def apply_number_formats(ws):
    """依欄位名稱套用 Excel number_format，保留真正數字型別。"""
    headers = get_header_map(ws)

    money_columns = {
        "台幣均價",
        "台幣現價",
        "台幣成本",
        "台幣市值",
        "台幣損益",
        "market_value_twd",
        "cost_value_twd",
        "profit_twd",
    }
    decimal_columns = {
        "股數",
        "原始均價",
        "原始現價",
        "原始損益",
        "shares",
        "cost",
        "price",
        "profit",
    }
    percent_columns = {
        "報酬率",
        "持倉比例",
        "return_rate",
        "weight",
    }

    for col_name, col_idx in headers.items():
        number_format = None

        if col_name in money_columns:
            number_format = '"NT$"#,##0'
        elif col_name in decimal_columns:
            number_format = '#,##0.00'
        elif col_name in percent_columns:
            number_format = '0.00%'

        if number_format:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = number_format

    # Summary 是 metric/value 形式，需要依 metric 決定 value 格式。
    if ws.title == EXCEL_SHEET_NAMES["summary"] and {"metric", "value"}.issubset(headers):
        metric_col = headers["metric"]
        value_col = headers["value"]
        money_metrics = {"總市值", "總成本", "總損益"}
        percent_metrics = {"總報酬率", "美股佔比", "台股佔比"}

        for row in range(2, ws.max_row + 1):
            metric = ws.cell(row=row, column=metric_col).value
            value_cell = ws.cell(row=row, column=value_col)

            if metric in money_metrics:
                value_cell.number_format = '"NT$"#,##0'
            elif metric in percent_metrics:
                value_cell.number_format = '0.00%'


def apply_profit_loss_colors(ws):
    """正數顯示綠色，負數顯示紅色。"""
    headers = get_header_map(ws)
    green_font = Font(color="FF00B050")
    red_font = Font(color="FFFF0000")

    target_columns = ["原始損益", "台幣損益", "報酬率"]

    for col_name in target_columns:
        col_idx = headers.get(col_name)
        if not col_idx:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = green_font
                elif cell.value < 0:
                    cell.font = red_font

    if ws.title == EXCEL_SHEET_NAMES["summary"] and {"metric", "value"}.issubset(headers):
        metric_col = headers["metric"]
        value_col = headers["value"]
        for row in range(2, ws.max_row + 1):
            metric = ws.cell(row=row, column=metric_col).value
            cell = ws.cell(row=row, column=value_col)
            if metric in {"總損益", "總報酬率"} and isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = green_font
                elif cell.value < 0:
                    cell.font = red_font


def auto_adjust_column_widths(ws):
    """依內容長度調整欄寬，避免文字擠在一起。"""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 32)


def format_worksheet_header(ws):
    """套用第一列表頭樣式。"""
    header_fill = PatternFill("solid", fgColor="14232B")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font


def highlight_price_status_rows(ws):
    """價格狀態不是價格正常時，整列用淡黃色標示。"""
    headers = get_header_map(ws)
    status_col = headers.get("價格狀態")
    if not status_col:
        return

    warning_fill = PatternFill("solid", fgColor="FFF2CC")

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value
        if status and status != "價格正常":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = warning_fill


def add_excel_table_and_filter(ws, sheet_name):
    """替每個 sheet 建立 Excel Table 與篩選器。"""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table_name = "Table_" + "".join(ch if ch.isalnum() else "_" for ch in sheet_name)

    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def create_png_download(fig):
    """把 matplotlib dashboard 存成 PNG bytes。"""
    output = BytesIO()
    fig.savefig(output, format="png", dpi=300, bbox_inches="tight", facecolor=BACKGROUND_COLOR)
    return output.getvalue()


def render_download_button(label, data, file_name, mime, container=st):
    """統一建立下載按鈕，避免同樣的參數寫法重複出現。"""
    container.download_button(
        label=label,
        data=data,
        file_name=file_name,
        mime=mime,
    )


def format_twd(value):
    """把數字格式化成台幣金額，例如 NT$1,234,567。"""
    if pd.isna(value):
        return "N/A"
    return f"NT${value:,.0f}"


def format_decimal(value, digits=2):
    """把一般數字格式化成固定小數位，例如 1,234.56。"""
    if pd.isna(value):
        return "N/A"
    return f"{value:,.{digits}f}"


def format_percent(value):
    """把小數格式化成百分比，例如 0.1234 會變成 12.34%。"""
    if pd.isna(value):
        return "N/A"
    return f"{value:.2%}"


def create_market_allocation(df):
    """依市場彙總台幣市值，建立台股 / 美股配置表。"""
    allocation = (
        df.groupby("market", as_index=False)["market_value_twd"]
        .sum()
        .sort_values("market_value_twd", ascending=False)
    )
    total_value = allocation["market_value_twd"].sum()
    allocation["weight"] = allocation["market_value_twd"] / total_value if total_value else 0
    allocation["market_name"] = allocation["market"].map(MARKET_NAME_MAP)
    return allocation


def create_category_allocation(df):
    """依 category 彙總台幣市值，建立類別配置分析資料。"""
    allocation = (
        df.groupby("category", as_index=False)["market_value_twd"]
        .sum()
        .sort_values("market_value_twd", ascending=False)
    )
    total_value = allocation["market_value_twd"].sum()
    allocation["weight"] = allocation["market_value_twd"] / total_value if total_value else 0
    return allocation


def create_category_display_table(category_allocation):
    """把類別配置表轉成適合畫面閱讀的格式，保持數字型態。"""
    display_df = category_allocation.copy()
    display_df = display_df.reset_index(drop=True)
    
    # 保持數字型態，不轉換成字串
    return display_df.rename(
        columns={
            "category": "類別",
            "market_value_twd": "台幣市值",
            "weight": "持倉比例",
        }
    )


def plot_category_chart(category_allocation, category_top_n):
    """畫出 category 類別配置 donut chart，使用 Top N + 其他。"""
    fig, ax = plt.subplots(figsize=(9, 5.6))
    fig.patch.set_facecolor(BACKGROUND_COLOR)
    ax.set_facecolor(BACKGROUND_COLOR)

    if category_allocation.empty:
        ax.text(0.5, 0.5, "無類別資料", ha="center", va="center", color="white")
        ax.axis("off")
        return fig

    chart_df = make_top_n_summary(
        data=category_allocation,
        name_col="category",
        value_col="market_value_twd",
        top_n=category_top_n,
    )

    plot_donut_on_ax(
        ax=ax,
        summary=chart_df,
        name_col="category",
        value_col="market_value_twd",
        title=f"類別配置前 {category_top_n} 大 + 其他",
        center_text="類別",
        label_size=10,
        pct_size=9,
        center_size=15,
        title_size=14,
        title_y=1.08,
        radius=0.82,
    )

    fig.tight_layout()
    return fig


def plot_category_allocation(category_allocation, category_top_n):
    return plot_category_chart(category_allocation, category_top_n)


def create_risk_flags(df):
    """建立風險提醒資料，例如高集中持股或股價抓取失敗。"""
    flags = []

    for _, row in df[df["weight"] > 0.3].iterrows():
        flags.append(
            {
                "type": "高集中持股",
                "symbol": row["symbol"],
                "detail": f"單一持股比例為 {format_percent(row['weight'])}，超過 30%。",
            }
        )

    for _, row in df[df["price_status"] != "價格正常"].iterrows():
        flags.append(
            {
                "type": "價格抓取失敗",
                "symbol": row["symbol"],
                "detail": "此股票暫用成本估算市值與損益。",
            }
        )

    if not flags:
        flags.append({"type": "無", "symbol": "", "detail": "目前沒有需要提醒的風險旗標。"})

    return pd.DataFrame(flags)


def create_summary_metrics(df):
    """計算摘要卡片需要用到的投資組合指標。"""
    total_value = df["market_value_twd"].sum()
    total_cost = df["cost_value_twd"].sum()
    total_profit = df["profit_twd"].sum()
    total_return = total_profit / total_cost if total_cost else 0

    # 最大持股用持倉比例最高的那一筆資料。
    largest_position = df.sort_values("weight", ascending=False).iloc[0]
    largest_holding = (
        f"{largest_position['symbol']} ({format_percent(largest_position['weight'])})"
    )

    if total_value:
        market_weights = df.groupby("market")["market_value_twd"].sum() / total_value
        us_weight = market_weights.get("US", 0)
        tw_weight = market_weights.get("TW", 0)
    else:
        us_weight = 0
        tw_weight = 0

    return {
        "total_value": total_value,
        "total_cost": total_cost,
        "total_profit": total_profit,
        "total_return": total_return,
        "largest_holding": largest_holding,
        "us_weight": us_weight,
        "tw_weight": tw_weight,
    }


def create_display_table(df, sort_option):
    """建立畫面上使用的中文報表，保持數字型態便於排序。
    
    數字欄位保持 float/numeric 種點，不轉換成字串。
    顯示格式化戲碼由 st.dataframe 的 column_config 來處理。
    """
    # 先用完整資料排序，排序完成後才挑出要顯示的欄位。
    sorted_df = sort_holdings_for_display(df, sort_option)

    display_df = sorted_df[DISPLAY_COLUMNS].copy()

    # 使用 reset_index 移除原始 index，st.dataframe 用 hide_index 隱藏
    display_df = display_df.reset_index(drop=True)

    # 市場代碼轉成中文，但保持其他數字欄位型態
    display_df["market"] = display_df["market"].map(MARKET_NAME_MAP)

    # 重命名欄位為中文，保持數字型態
    display_df = display_df.rename(columns=DISPLAY_COLUMN_NAMES)

    return display_df


def get_profit_loss_text_color(value):
    """依照正負數回傳表格文字顏色：正數綠色、負數紅色。"""
    if pd.isna(value):
        return ""
    if value > 0:
        return "color: #00B050; font-weight: 700;"
    if value < 0:
        return "color: #FF6B6B; font-weight: 700;"
    return ""


def style_holdings_table(display_table):
    """替完整持倉報表加上損益紅綠色，不改變原本數字資料。"""
    color_columns = [
        col
        for col in ["原始損益", "台幣損益", "報酬率"]
        if col in display_table.columns
    ]
    if not color_columns:
        return display_table

    # pandas 新版使用 Styler.map，避免舊的 applymap 相容性問題。
    return display_table.style.map(get_profit_loss_text_color, subset=color_columns)


def sort_holdings_for_display(df, sort_option):
    """依使用者選擇的排序方式排列持倉報表。"""
    sort_rules = {
        "照字母開頭": ("symbol", True),
        "報酬率": ("return_rate", False),
        "台幣損益": ("profit_twd", False),
        "持倉比重": ("market_value_twd", False),
    }
    sort_column, ascending = sort_rules.get(sort_option, ("market_value_twd", False))
    return df.sort_values(sort_column, ascending=ascending)


def apply_category_edits(report, edited_display_table, sort_option):
    """把持倉報表中編輯後的「類別」欄位寫回原始 report。"""
    updated_report = report.copy()

    # data_editor 回傳的資料列順序，會跟畫面上排序後的報表一致。
    sorted_index = sort_holdings_for_display(updated_report, sort_option).index.tolist()

    for row_position, original_index in enumerate(sorted_index):
        selected_category = edited_display_table.iloc[row_position]["類別"]
        updated_report.at[original_index, "category"] = selected_category

    return updated_report


def render_summary_metrics(summary):
    """用較不擁擠的方式顯示摘要卡片，手機版也比較容易閱讀。"""
    metric_col_1, metric_col_2 = st.columns(2)
    metric_col_1.metric("總市值", format_twd(summary["total_value"]))
    metric_col_2.metric("總成本", format_twd(summary["total_cost"]))

    metric_col_3, metric_col_4 = st.columns(2)
    metric_col_3.metric("總損益", format_twd(summary["total_profit"]))
    metric_col_4.metric("總報酬率", format_percent(summary["total_return"]))

    metric_col_5, metric_col_6, metric_col_7 = st.columns(3)
    metric_col_5.metric("最大持股", summary["largest_holding"])
    metric_col_6.metric("美股佔比", format_percent(summary["us_weight"]))
    metric_col_7.metric("台股佔比", format_percent(summary["tw_weight"]))


def render_risk_messages(report):
    """顯示股價抓取失敗與高集中持股提醒。"""
    failed_price_symbols = report[report["price_status"] != "價格正常"]["symbol"].tolist()
    if failed_price_symbols:
        st.warning(
            "以下股票無法取得最新股價，系統已暫用成本估算市值與損益："
            + "、".join(failed_price_symbols)
        )

    concentrated_positions = report[report["weight"] > 0.3]["symbol"].tolist()
    if concentrated_positions:
        symbols_text = "、".join(concentrated_positions)
        st.warning(f"注意：以下持股超過 30%，投資組合可能過度集中：{symbols_text}")
    else:
        st.success("目前沒有單一持股超過 30%。")


def render_overview_tab(report, summary, usd_twd, top_n):
    """總覽 tab：摘要卡片、提醒與 Dashboard 圖。"""
    st.subheader("投資組合摘要")
    render_summary_metrics(summary)
    render_portfolio_health(report)
    render_risk_messages(report)
    st.caption(f"USD/TWD 匯率：{usd_twd:.4f}")

    st.subheader("投資組合總覽")
    fig = plot_dashboard(report, top_n)
    st.pyplot(fig, use_container_width=True)
    render_download_button(
        label="下載 Dashboard PNG",
        data=create_png_download(fig),
        file_name="portfolio_dashboard.png",
        mime="image/png",
    )

    if len(report) > 10:
        top_holdings_fig = plot_top_holdings_bar(report, top_n)
        st.subheader(f"前 {top_n} 大持股橫向長條圖")
        st.pyplot(top_holdings_fig, use_container_width=True)
        render_download_button(
            label="下載前大持股長條圖 PNG",
            data=create_png_download(top_holdings_fig),
            file_name="top_holdings_bar.png",
            mime="image/png",
        )
        plt.close(top_holdings_fig)

    return fig


def render_holdings_tab(report, sort_option):
    """持倉 tab：在報表中調整類別，並顯示有條件色彩的完整報表。"""
    st.subheader("持倉報表")
    st.write("可以先新增自訂類別，再直接在下方完整持倉報表的「類別」欄位調整分類。")

    custom_category = st.text_input(
        "新增自訂 category",
        placeholder="例如：高股息ETF、債券、現金部位",
    )
    editable_category_options = update_category_options(report, custom_category)

    display_table = create_display_table(report, sort_option=sort_option)
    locked_columns = [col for col in display_table.columns if col != "類別"]

    st.markdown("#### 完整持倉報表")
    
    # 定義 column_config 來格式化數字顯示，同時保持原始數字型態便於排序
    column_config = {
        "類別": st.column_config.SelectboxColumn(
            "類別",
            options=editable_category_options,
            required=True,
        ),
        "股數": st.column_config.NumberColumn(format="%.2f"),
        "原始均價": st.column_config.NumberColumn(format="%.2f"),
        "原始現價": st.column_config.NumberColumn(format="%.2f"),
        "原始損益": st.column_config.NumberColumn(format="%.0f"),
        "台幣現值": st.column_config.NumberColumn(format="NT$%.0f"),
        "台幣損益": st.column_config.NumberColumn(format="NT$%.0f"),
        "報酬率": st.column_config.NumberColumn(format="percent"),
        "持倉比例": st.column_config.NumberColumn(format="percent"),
    }
    
    edited_display_table = st.data_editor(
        style_holdings_table(display_table),
        use_container_width=True,
        disabled=locked_columns,
        column_config=column_config,
        hide_index=True,
    )
    updated_report = apply_category_edits(report, edited_display_table, sort_option)

    if (updated_report["category"] == "未分類").any():
        st.warning("有持股尚未分類，建議補上 category，類別配置圖會更準確。")
    else:
        st.success("所有持股都已完成分類。")

    return updated_report


def render_category_tab(category_allocation, category_top_n):
    """類別 tab：類別配置表、donut chart 與簡短說明。"""
    st.subheader("類別配置")
    st.write("類別配置會依台幣市值由大到小排序，圖表會顯示前幾大，其餘合併成「其他」。")
    
    category_table = create_category_display_table(category_allocation)
    
    # 使用 column_config 來格式化顯示，保持底層數字型態
    column_config = {
        "台幣市值": st.column_config.NumberColumn(format="NT$%.0f"),
        "持倉比例": st.column_config.NumberColumn(format="percent"),
    }
    
    st.dataframe(
        category_table,
        column_config=column_config,
        use_container_width=True,
        hide_index=True,
    )

    category_fig = plot_category_allocation(category_allocation, category_top_n)
    st.pyplot(category_fig, use_container_width=True)

    return category_fig


def render_excel_report_download(report, summary, market_allocation, category_allocation, risk_flags):
    """在持倉報表下方提供 Excel 多 sheet 報表下載。"""
    excel_bytes = create_multi_sheet_excel(
        report,
        summary,
        market_allocation,
        category_allocation,
        risk_flags,
    )
    render_download_button(
        label="下載完整 Excel 報表",
        data=excel_bytes,
        file_name=f"portfolio_report_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def render_page_intro():
    """顯示頁面標題、說明與使用步驟。"""
    st.title("個人投資組合分析器")
    st.write(
        "這個工具可以上傳持倉 Excel（.xlsx/.xls），自動抓取股價，統一換算成台幣，"
        "並產生持倉報表與投資組合圖表。"
    )

    st.subheader("使用步驟")
    st.markdown(
        """
        1. Step 1：準備持倉 Excel 檔
        2. Step 2：上傳 Excel
        3. Step 3：系統自動讀取表格、抓取股價並換算台幣
        4. Step 4：查看報表與圖表
        5. Step 5：下載 Excel 報表或 PNG 圖片
        """
    )


def render_sidebar_controls():
    """顯示側邊欄設定，並回傳使用者選擇的控制值。"""
    top_n = st.sidebar.slider("圖表顯示前幾大", min_value=1, max_value=20, value=7)
    category_top_n = st.sidebar.slider(
        "類別配置顯示前幾大",
        min_value=1,
        max_value=20,
        value=7,
    )
    sort_option = st.sidebar.radio(
        "持倉報表排序",
        ["持倉比重", "報酬率", "台幣損益", "照字母開頭"],
    )
    use_excel_price = st.sidebar.checkbox(
        "若 Excel 有 price 欄位，優先使用該欄位作為現價",
        value=False,
    )
    debug_mode = st.sidebar.checkbox("顯示除錯資訊", value=False)

    st.sidebar.subheader(
        "Excel 可以有標題列或空白列，系統會自動偵測欄位名稱列。"
    )
    st.sidebar.markdown(
        """
        - symbol：股票代號，例如 NVDA、TSLA、0050
        - market：市場，只能填 US 或 TW
        - shares：股數，可填小數
        - cost：平均成本
        - category：分類，可留空，系統會自動填入未分類
        - price：可選欄位，若有可當作現價使用
        """
    )
    sample_data, sample_mime, sample_name = create_sample_excel_bytes()
    render_download_button(
        label="下載範例 Excel",
        data=sample_data,
        file_name=sample_name,
        mime=sample_mime,
        container=st.sidebar,
    )

    return top_n, category_top_n, sort_option, use_excel_price, debug_mode


def main():
    st.set_page_config(
        page_title="Personal Portfolio Analyzer",
        page_icon="📊",
        layout="wide",
    )

    st.markdown(
        f"""
        <style>
        .stApp {{
            background-color: {BACKGROUND_COLOR};
            color: white;
        }}
        .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {{
            color: white;
        }}
        .stApp p, .stApp div, .stApp label, .stApp span {{
            color: white;
        }}
        [data-testid="stSidebar"] {{
            background-color: #0f1b22;
        }}
        [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {{
            color: white;
        }}
        [data-testid="stSidebar"] p, [data-testid="stSidebar"] label, [data-testid="stSidebar"] div {{
            color: #d0d0d0;
        }}
        div[data-testid="stMetric"] {{
            background-color: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 8px;
            padding: 12px;
        }}
        div[data-testid="stMetric"] label {{
            color: #b0b0b0;
        }}
        div[data-testid="stMetric"] > div {{
            color: white;
        }}
        @media (max-width: 768px) {{
            .block-container {{
                padding-left: 1rem;
                padding-right: 1rem;
            }}
            div[data-testid="stMetric"] {{
                padding: 10px;
            }}
            .stDownloadButton button {{
                width: 100%;
            }}
        }}
        .stTabs [data-baseweb="tab-list"] button {{
            color: #d0d0d0;
        }}
        .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {{
            color: white;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    render_page_intro()
    top_n, category_top_n, sort_option, use_excel_price, debug_mode = render_sidebar_controls()

    upload_col, sample_col = st.columns([3, 1])
    uploaded_file = upload_col.file_uploader("上傳持倉 Excel 檔", type=["xlsx", "xls"])
    use_sample = sample_col.checkbox("使用範例資料試玩")

    if use_sample:
        portfolio_df, inspection = load_sample_portfolio()
        st.success("已載入範例資料，可直接開始分析。")
    elif uploaded_file is not None:
        portfolio_df, inspection = load_portfolio(uploaded_file, debug_mode=debug_mode)
    else:
        st.info("請先上傳持倉 Excel 檔，或勾選「使用範例資料試玩」。")
        return

    if not show_data_check_results(inspection):
        return

    try:
        with st.spinner("正在抓最新股價與 USD/TWD 匯率..."):
            usd_twd = get_usd_twd_rate()
            report = calculate_portfolio(portfolio_df, usd_twd, use_excel_price=use_excel_price)

        summary = create_summary_metrics(report)
        overview_tab, holdings_tab, category_tab = st.tabs(["總覽", "持倉", "類別"])

        with overview_tab:
            fig = render_overview_tab(report, summary, usd_twd, top_n)

        with holdings_tab:
            report = render_holdings_tab(report, sort_option)
            market_allocation = create_market_allocation(report)
            category_allocation = create_category_allocation(report)
            risk_flags = create_risk_flags(report)
            render_excel_report_download(
                report,
                summary,
                market_allocation,
                category_allocation,
                risk_flags,
            )

        with category_tab:
            category_fig = render_category_tab(category_allocation, category_top_n)

        plt.close(fig)
        plt.close(category_fig)

    except Exception as exc:
        st.error(str(exc))


if __name__ == "__main__":
    main()
